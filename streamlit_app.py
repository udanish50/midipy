from __future__ import annotations

import hashlib
import io
import json
import math
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from midipy.midi_analysis import midiInfo
from midipy.midi_parser import parser, parser_segments
from midipy.midi_reader import readmidi
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# APP CONFIGURATION
# =============================================================================

st.set_page_config(
    page_title="MidiPy Analysis Studio",
    page_icon="🎵",
    layout="wide",
    initial_sidebar_state="collapsed",
)

DEFAULT_UE_KEYS = [
    24, 25, 26, 27, 28, 30, 31, 32, 34, 35,
    37, 38, 39, 40, 41, 42, 43, 46, 47, 48,
    49, 51, 52, 53, 55, 59, 66, 67, 71, 78, 79,
]

METRIC_GROUPS = {
    "Note counts": [
        "Total_Counts",
        "UE_Counts",
        "LF_Counts",
        "RF_Counts",
    ],
    "Velocity": [
        "Avg_Velocity",
        "UE_Velocity",
        "LF_Velocity",
        "RF_Velocity",
    ],
    "Asynchrony": [
        "Avg_Async",
        "UE_Async",
        "LF_Async",
        "RF_Async",
    ],
}

AVAILABLE_METRICS = [
    metric
    for metrics in METRIC_GROUPS.values()
    for metric in metrics
]

METRIC_LABELS = {
    "Name": "Session",
    "Total_Counts": "Total notes",
    "UE_Counts": "Upper-extremity notes",
    "LF_Counts": "Left-foot notes",
    "RF_Counts": "Right-foot notes",
    "Avg_Velocity": "Overall velocity",
    "UE_Velocity": "Upper-extremity velocity",
    "LF_Velocity": "Left-foot velocity",
    "RF_Velocity": "Right-foot velocity",
    "Avg_Async": "Overall asynchrony",
    "UE_Async": "Upper-extremity asynchrony",
    "LF_Async": "Left-foot asynchrony",
    "RF_Async": "Right-foot asynchrony",
    "Segment_Number": "Segment",
}

METRIC_HELP = {
    "Total_Counts": "Total number of MIDI notes detected.",
    "UE_Counts": "Notes whose MIDI values match the upper-extremity mapping.",
    "LF_Counts": "Notes whose MIDI value matches the left-foot mapping.",
    "RF_Counts": "Notes whose MIDI value matches the right-foot mapping.",
    "Avg_Velocity": "Mean note velocity with standard deviation in parentheses.",
    "UE_Velocity": "Upper-extremity mean velocity with standard deviation.",
    "LF_Velocity": "Left-foot mean velocity with standard deviation.",
    "RF_Velocity": "Right-foot mean velocity with standard deviation.",
    "Avg_Async": "Mean overall timing asynchrony with standard deviation.",
    "UE_Async": "Upper-extremity timing asynchrony.",
    "LF_Async": "Left-foot timing asynchrony.",
    "RF_Async": "Right-foot timing asynchrony.",
}

COUNT_METRICS = [
    "Total_Counts",
    "UE_Counts",
    "LF_Counts",
    "RF_Counts",
]


FOOT_METRICS = [
    "LF_Counts",
    "RF_Counts",
    "LF_Velocity",
    "RF_Velocity",
    "LF_Async",
    "RF_Async",
]

NON_FOOT_METRICS = [
    metric
    for metric in AVAILABLE_METRICS
    if metric not in FOOT_METRICS
]

NON_FOOT_COUNT_METRICS = [
    metric
    for metric in COUNT_METRICS
    if metric not in {"LF_Counts", "RF_Counts"}
]


# A new cycle creates fresh widget keys so "Start new analysis" truly resets
# uploaded files, mappings, analysis choices, and displayed results.
if "analysis_cycle" not in st.session_state:
    st.session_state["analysis_cycle"] = 0

analysis_cycle = int(st.session_state["analysis_cycle"])



# =============================================================================
# HUMAN-CENTRED LIGHT VISUAL SYSTEM
# =============================================================================

st.markdown(
    """
    <style>
    :root {
        --mp-primary: #3157d5;
        --mp-primary-dark: #2340aa;
        --mp-primary-soft: #eef3ff;
        --mp-text: #172033;
        --mp-muted: #58667c;
        --mp-border: #dfe5ef;
        --mp-surface: #ffffff;
        --mp-background: #f6f8fc;
        --mp-success: #166b44;
        --mp-success-soft: #edf8f2;
        --mp-warning: #8a5200;
        --mp-warning-soft: #fff7e8;
        --mp-danger: #a62d27;
        --mp-danger-soft: #fff0ef;
        --mp-radius: 16px;
        --mp-shadow: 0 10px 26px rgba(31, 47, 86, 0.07);
    }

    html {
        color-scheme: light !important;
        scroll-behavior: smooth;
    }

    body,
    .stApp,
    [data-testid="stAppViewContainer"] {
        background: var(--mp-background) !important;
        color: var(--mp-text) !important;
    }

    [data-testid="stHeader"] {
        background: rgba(246, 248, 252, 0.94) !important;
        border-bottom: 1px solid rgba(223, 229, 239, 0.88);
        backdrop-filter: blur(12px);
    }

    [data-testid="stSidebar"],
    [data-testid="collapsedControl"] {
        display: none !important;
    }

    .block-container {
        max-width: 1160px;
        padding-top: 1.25rem;
        padding-bottom: 5.2rem;
    }

    h1, h2, h3 {
        color: var(--mp-text);
        letter-spacing: -0.025em;
    }

    p, label, li {
        line-height: 1.5;
    }

    .mp-header-title {
        margin: 0;
        font-size: clamp(1.75rem, 4vw, 2.35rem);
        line-height: 1.12;
    }

    .mp-header-subtitle {
        margin: 0.35rem 0 0;
        color: var(--mp-muted);
        font-size: 0.98rem;
    }

    .mp-section-title {
        margin: 1.3rem 0 0.15rem;
        font-size: 1.35rem;
    }

    .mp-section-copy {
        margin: 0 0 0.7rem;
        color: var(--mp-muted);
        font-size: 0.94rem;
    }

    .mp-progress {
        display: grid;
        grid-template-columns: auto 1fr auto 1fr auto 1fr auto;
        align-items: center;
        gap: 0.55rem;
        margin: 1rem 0 1.25rem;
        padding: 0.72rem 0.85rem;
        border: 1px solid var(--mp-border);
        border-radius: 14px;
        background: var(--mp-surface);
        box-shadow: 0 5px 16px rgba(31, 47, 86, 0.04);
    }

    .mp-progress-step {
        display: flex;
        align-items: center;
        gap: 0.45rem;
        min-width: 0;
        color: #7b8699;
        font-size: 0.85rem;
        font-weight: 700;
        white-space: nowrap;
    }

    .mp-progress-step span {
        display: inline-grid;
        width: 1.7rem;
        height: 1.7rem;
        flex: 0 0 1.7rem;
        place-items: center;
        border: 1px solid #cad2df;
        border-radius: 50%;
        background: #f7f8fb;
        color: #657187;
        font-size: 0.78rem;
    }

    .mp-progress-step.is-active {
        color: var(--mp-primary-dark);
    }

    .mp-progress-step.is-active span {
        border-color: var(--mp-primary);
        background: var(--mp-primary);
        color: #ffffff;
        box-shadow: 0 0 0 4px rgba(49, 87, 213, 0.12);
    }

    .mp-progress-step.is-complete {
        color: #315e49;
    }

    .mp-progress-step.is-complete span {
        border-color: #8dc5aa;
        background: var(--mp-success-soft);
        color: var(--mp-success);
    }

    .mp-progress-line {
        height: 2px;
        background: #dfe4ec;
    }

    .mp-progress-line.is-complete {
        background: #8dc5aa;
    }

    .mp-upload-summary {
        display: flex;
        flex-wrap: wrap;
        align-items: center;
        gap: 0.75rem 1.25rem;
        margin-top: 0.8rem;
        padding: 0.8rem 0.9rem;
        border: 1px solid var(--mp-border);
        border-radius: 12px;
        background: #fafbfd;
        font-size: 0.9rem;
    }

    .mp-upload-summary strong {
        color: var(--mp-text);
    }

    .mp-ready {
        color: var(--mp-success);
    }

    .mp-attention {
        color: var(--mp-warning);
    }

    .mp-card-heading {
        margin: 0 0 0.2rem;
        font-size: 1.05rem;
    }

    .mp-card-copy {
        margin: 0 0 0.8rem;
        color: var(--mp-muted);
        font-size: 0.88rem;
    }

    .mp-inline-warning {
        margin-top: 0.6rem;
        padding: 0.72rem 0.82rem;
        border: 1px solid #eed09b;
        border-radius: 11px;
        background: var(--mp-warning-soft);
        color: #684000;
        font-size: 0.88rem;
    }

    .mp-inline-ok {
        margin-top: 0.6rem;
        padding: 0.68rem 0.8rem;
        border: 1px solid #b7dcc8;
        border-radius: 11px;
        background: var(--mp-success-soft);
        color: #24583f;
        font-size: 0.87rem;
    }

    .mp-results-banner {
        display: flex;
        flex-wrap: wrap;
        align-items: center;
        gap: 0.55rem;
        padding: 0.88rem 1rem;
        border: 1px solid #b7dcc8;
        border-radius: 13px;
        background: var(--mp-success-soft);
        color: #24583f;
        font-weight: 700;
    }

    .mp-muted-note {
        color: var(--mp-muted);
        font-size: 0.86rem;
    }

    .mp-footer {
        margin-top: 2.2rem;
        padding-top: 1rem;
        border-top: 1px solid var(--mp-border);
        color: #6a7689;
        font-size: 0.8rem;
        text-align: center;
    }

    div[data-testid="stVerticalBlockBorderWrapper"] {
        border-color: var(--mp-border) !important;
        border-radius: var(--mp-radius) !important;
        background: var(--mp-surface) !important;
        box-shadow: var(--mp-shadow);
    }

    div[data-testid="stFileUploaderDropzone"] {
        min-height: 142px;
        border: 2px dashed #aebce0 !important;
        border-radius: 14px !important;
        background: #f9fbff !important;
        transition: 150ms ease;
    }

    div[data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--mp-primary) !important;
        background: var(--mp-primary-soft) !important;
    }

    div.stButton > button,
    div.stDownloadButton > button {
        min-height: 44px;
        border-radius: 11px;
        font-weight: 720;
        transition: transform 130ms ease, box-shadow 130ms ease;
    }

    div.stButton > button:hover,
    div.stDownloadButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 7px 18px rgba(35, 64, 170, 0.13);
    }

    div.stButton > button:focus-visible,
    div.stDownloadButton > button:focus-visible,
    input:focus-visible,
    textarea:focus-visible,
    [role="combobox"]:focus-visible {
        outline: 3px solid rgba(49, 87, 213, 0.30) !important;
        outline-offset: 2px;
    }

    button[kind="primary"] {
        border-color: var(--mp-primary) !important;
        background: linear-gradient(135deg, var(--mp-primary), var(--mp-primary-dark)) !important;
        color: #ffffff !important;
    }

    button:disabled {
        opacity: 1 !important;
        border-color: #d5dbe5 !important;
        background: #e8ecf3 !important;
        color: #687487 !important;
    }

    [data-testid="stRadio"] div[role="radiogroup"] {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        gap: 0.55rem;
    }

    [data-testid="stRadio"] div[role="radiogroup"] > label {
        min-height: 54px;
        margin: 0;
        padding: 0.72rem 0.75rem;
        align-items: flex-start;
        border: 1px solid var(--mp-border);
        border-radius: 12px;
        background: #ffffff;
        transition: border-color 130ms ease, background 130ms ease, box-shadow 130ms ease;
    }

    [data-testid="stRadio"] div[role="radiogroup"] > label:hover {
        border-color: #a8b8e8;
        background: #fafbff;
    }

    [data-testid="stRadio"] div[role="radiogroup"] > label:has(input:checked) {
        border-color: var(--mp-primary);
        background: var(--mp-primary-soft);
        box-shadow: 0 0 0 2px rgba(49, 87, 213, 0.08);
    }

    [data-testid="stMultiSelect"] [data-baseweb="tag"] {
        min-height: 30px;
        border-radius: 8px;
        background: var(--mp-primary-soft);
        color: var(--mp-primary-dark);
    }

    [data-testid="stMetric"] {
        min-height: 104px;
        padding: 0.9rem;
        border: 1px solid var(--mp-border);
        border-radius: 13px;
        background: #ffffff;
        box-shadow: 0 5px 16px rgba(31, 47, 86, 0.045);
    }

    /* Results navigation: separate, spacious tab cards */
    [data-baseweb="tab-list"] {
        display: grid !important;
        grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        gap: 0.8rem !important;
        width: 100%;
        max-width: 820px;
        margin: 0 0 1rem;
        padding: 0 !important;
        border: 0 !important;
        background: transparent !important;
    }

    [data-baseweb="tab"] {
        min-height: 50px;
        margin: 0 !important;
        padding: 0.72rem 1rem !important;
        justify-content: center;
        border: 1px solid #dbe2ed !important;
        border-radius: 12px !important;
        background: #ffffff !important;
        color: #536177 !important;
        font-weight: 720 !important;
        box-shadow: 0 4px 12px rgba(31, 47, 86, 0.045);
        transition:
            border-color 140ms ease,
            background 140ms ease,
            color 140ms ease,
            transform 140ms ease,
            box-shadow 140ms ease;
    }

    [data-baseweb="tab"] p,
    [data-baseweb="tab"] span {
        margin: 0 !important;
        color: inherit !important;
        font-weight: inherit !important;
        white-space: nowrap;
    }

    [data-baseweb="tab"]:hover {
        border-color: #9eb0e5 !important;
        background: #f7f9ff !important;
        color: #2948b8 !important;
        transform: translateY(-1px);
        box-shadow: 0 7px 17px rgba(31, 47, 86, 0.08);
    }

    [data-baseweb="tab"]:focus-visible {
        outline: 3px solid rgba(49, 87, 213, 0.28) !important;
        outline-offset: 2px;
    }

    [aria-selected="true"][data-baseweb="tab"] {
        border-color: #3157d5 !important;
        background: linear-gradient(135deg, #3157d5, #2545b6) !important;
        color: #ffffff !important;
        box-shadow: 0 8px 20px rgba(49, 87, 213, 0.22) !important;
        transform: translateY(-1px);
    }

    [aria-selected="true"][data-baseweb="tab"]:hover {
        background: linear-gradient(135deg, #2d51cb, #213fa9) !important;
        color: #ffffff !important;
    }

    [data-baseweb="tab-highlight"],
    [data-baseweb="tab-border"] {
        display: none !important;
    }

    [data-testid="stDataFrame"] {
        overflow: hidden;
        border: 1px solid var(--mp-border);
        border-radius: 12px;
    }

    /* The :has selector keeps the real Streamlit button in a sticky card
       without creating a fragile custom JavaScript component. */
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.mp-sticky-marker) {
        position: sticky;
        bottom: 0.65rem;
        z-index: 999;
        border-color: #cbd6f3 !important;
        background: rgba(255, 255, 255, 0.97) !important;
        box-shadow: 0 14px 34px rgba(31, 47, 86, 0.17);
        backdrop-filter: blur(10px);
    }

    @media (max-width: 760px) {
        .block-container {
            padding-top: 0.85rem;
            padding-inline: 0.75rem;
        }

        .mp-progress {
            grid-template-columns: 1fr;
            gap: 0.35rem;
        }

        .mp-progress-line {
            display: none;
        }

        [data-testid="stRadio"] div[role="radiogroup"] {
            grid-template-columns: 1fr;
        }

        [data-baseweb="tab-list"] {
            grid-template-columns: 1fr;
            gap: 0.55rem !important;
            width: 100%;
        }

        [data-baseweb="tab"] {
            justify-content: flex-start;
            min-height: 46px;
        }
    }

    @media (prefers-reduced-motion: reduce) {
        html {
            scroll-behavior: auto;
        }

        *,
        *::before,
        *::after {
            animation-duration: 0.01ms !important;
            animation-iteration-count: 1 !important;
            transition-duration: 0.01ms !important;
        }
    }
    
    /* Comparison mode */
    .mp-analysis-mode-note {
        margin-top: 0.55rem;
        padding: 0.72rem 0.82rem;
        border: 1px solid #d9e2f5;
        border-radius: 11px;
        background: #f7f9ff;
        color: #485873;
        font-size: 0.88rem;
    }

    .mp-role-heading {
        display: flex;
        align-items: center;
        gap: 0.55rem;
        margin-bottom: 0.2rem;
    }

    .mp-role-badge {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        min-width: 2rem;
        height: 2rem;
        padding: 0 0.62rem;
        border-radius: 999px;
        font-size: 0.76rem;
        font-weight: 800;
        letter-spacing: 0.025em;
    }

    .mp-role-badge.therapist {
        border: 1px solid #c7d4fb;
        background: #eef3ff;
        color: #294bbf;
    }

    .mp-role-badge.participant {
        border: 1px solid #b9dfd9;
        background: #edf9f6;
        color: #146b61;
    }

    .mp-role-title {
        margin: 0;
        color: #172033;
        font-size: 1.05rem;
        font-weight: 780;
    }

    .mp-role-copy {
        margin: 0.1rem 0 0.75rem;
        color: #5b6880;
        font-size: 0.86rem;
    }

    .mp-comparison-callout {
        padding: 0.82rem 0.92rem;
        border: 1px solid #d8e2f7;
        border-radius: 12px;
        background: linear-gradient(135deg, #f8faff, #f4fbfa);
        color: #40506c;
        font-size: 0.88rem;
    }

    .mp-comparison-banner {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 0.7rem;
        margin: 0.8rem 0 1rem;
    }

    .mp-comparison-banner > div {
        padding: 0.82rem 0.9rem;
        border: 1px solid #dfe5ef;
        border-radius: 12px;
        background: #ffffff;
        box-shadow: 0 5px 15px rgba(31, 47, 86, 0.045);
    }

    .mp-comparison-banner small {
        display: block;
        margin-bottom: 0.18rem;
        color: #69768a;
    }

    .mp-comparison-banner strong {
        color: #172033;
        font-size: 1rem;
    }

    .mp-neutral-note {
        margin-top: 0.55rem;
        color: #5b6880;
        font-size: 0.84rem;
    }

    @media (max-width: 760px) {
        .mp-comparison-banner {
            grid-template-columns: 1fr;
        }
    }

    </style>
    """,
    unsafe_allow_html=True,
)


# =============================================================================
# DATA AND VALIDATION HELPERS
# =============================================================================

def parse_note_list(text: str) -> list[int]:
    """Convert comma, semicolon, or whitespace-separated values into MIDI notes."""
    pieces = [item for item in re.split(r"[\s,;]+", text.strip()) if item]
    if not pieces:
        raise ValueError("Enter at least one upper-extremity MIDI note value.")

    values: list[int] = []
    for piece in pieces:
        try:
            value = int(piece)
        except ValueError as exc:
            raise ValueError(
                f'"{piece}" is not a whole MIDI note number.'
            ) from exc

        if not 0 <= value <= 127:
            raise ValueError(
                f"MIDI note {value} is outside the valid range of 0 to 127."
            )

        if value not in values:
            values.append(value)

    return values


def safe_filename(original_name: str, used_names: set[str]) -> str:
    """Create a unique lowercase .mid filename for MidiPy."""
    stem = Path(original_name).stem
    stem = re.sub(r"[^A-Za-z0-9_-]+", "_", stem).strip("_") or "midi_file"

    candidate = f"{stem}.mid"
    number = 2
    while candidate.lower() in used_names:
        candidate = f"{stem}_{number}.mid"
        number += 1

    used_names.add(candidate.lower())
    return candidate


def human_file_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024**2:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / 1024**2:.1f} MB"


def precheck_uploads(uploaded_files) -> tuple[list[dict[str, Any]], int, int]:
    """Perform a quick non-destructive check before the user starts analysis."""
    rows: list[dict[str, Any]] = []
    ready_count = 0
    total_size = 0

    for uploaded_file in uploaded_files or []:
        original_name = Path(uploaded_file.name).name
        data = uploaded_file.getvalue()
        size = len(data)
        total_size += size

        if original_name.startswith("."):
            status = "Needs attention"
            detail = "Hidden system file"
        elif size == 0:
            status = "Needs attention"
            detail = "Empty file"
        elif data[:4] != b"MThd":
            status = "Needs attention"
            detail = "Standard MIDI header is missing"
        else:
            status = "Ready"
            detail = "Basic file check passed"
            ready_count += 1

        rows.append(
            {
                "File": original_name,
                "Size": human_file_size(size),
                "Status": status,
                "Details": detail,
            }
        )

    return rows, ready_count, total_size


def validate_and_save_uploads(uploaded_files, folder: Path):
    """Save only genuine, MidiPy-readable files into a clean temporary folder."""
    valid_names: list[str] = []
    skipped: list[tuple[str, str]] = []
    used_names: set[str] = set()

    for uploaded_file in uploaded_files:
        original_name = Path(uploaded_file.name).name
        data = uploaded_file.getvalue()
        destination: Path | None = None

        try:
            if original_name.startswith("."):
                raise ValueError("Hidden system file")

            if not data:
                raise ValueError("The file is empty")

            if data[:4] != b"MThd":
                raise ValueError("The Standard MIDI MThd header is missing")

            destination_name = safe_filename(original_name, used_names)
            destination = folder / destination_name
            destination.write_bytes(data)

            midi = readmidi(str(destination))
            notes, _, tempos = midiInfo(midi, 0)

            if notes is None or len(notes) == 0:
                raise ValueError("No readable MIDI notes were found")

            if tempos is None or len(tempos) == 0:
                raise ValueError("No readable tempo information was found")

            valid_names.append(original_name)

        except Exception as error:
            if destination is not None and destination.exists():
                destination.unlink()
            skipped.append((original_name, str(error)))

    return valid_names, skipped


def upload_signature(uploaded_files) -> str:
    """Create a stable signature so stale results are never shown as current."""
    digest = hashlib.sha256()
    for uploaded_file in uploaded_files or []:
        data = uploaded_file.getvalue()
        digest.update(Path(uploaded_file.name).name.encode("utf-8", errors="ignore"))
        digest.update(str(len(data)).encode("ascii"))
        digest.update(hashlib.sha256(data).digest())
    return digest.hexdigest()


def settings_signature(settings: dict[str, Any]) -> str:
    payload = json.dumps(settings, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def mean_from_metric(value: Any) -> float | None:
    """Extract the leading mean from a value such as '45.37 (14.25)'."""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    match = re.search(r"[-+]?\d*\.?\d+", str(value))
    if not match:
        return None

    try:
        return float(match.group())
    except ValueError:
        return None


def display_dataframe(dataframe: pd.DataFrame) -> None:
    """Display a result table with human-readable column headings."""
    display_df = dataframe.rename(columns=METRIC_LABELS)

    column_config: dict[str, Any] = {}
    for source_name, display_name in METRIC_LABELS.items():
        if source_name not in dataframe.columns:
            continue

        help_text = METRIC_HELP.get(source_name)
        source_series = dataframe[source_name]

        if source_name in COUNT_METRICS or source_name == "Segment_Number":
            column_config[display_name] = st.column_config.NumberColumn(
                display_name,
                help=help_text,
                format="%.0f",
            )
        elif pd.api.types.is_numeric_dtype(source_series):
            column_config[display_name] = st.column_config.NumberColumn(
                display_name,
                help=help_text,
                format="%.2f",
            )
        else:
            column_config[display_name] = st.column_config.TextColumn(
                display_name,
                help=help_text,
            )

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config=column_config,
    )


def average_segment_rows(segment_df: pd.DataFrame) -> pd.DataFrame:
    """Average each segment index across files without using MidiPy's fragile path."""
    if "Name" not in segment_df.columns:
        return segment_df

    working = segment_df.copy()
    working["Segment_Number"] = pd.to_numeric(
        working["Name"]
        .astype(str)
        .str.extract(r"Segment\s+(\d+)", expand=False),
        errors="coerce",
    )
    working = working.dropna(subset=["Segment_Number"])

    if working.empty:
        return segment_df

    working["Segment_Number"] = working["Segment_Number"].astype(int)
    numeric_columns = [
        column
        for column in working.select_dtypes(include="number").columns
        if column != "Segment_Number"
    ]

    if not numeric_columns:
        return segment_df

    averaged = (
        working.groupby("Segment_Number", as_index=False)[numeric_columns]
        .mean()
        .sort_values("Segment_Number")
    )
    averaged.insert(
        0,
        "Name",
        "Segment " + averaged["Segment_Number"].astype(str),
    )

    ordered_columns = ["Name"] + [
        column
        for column in segment_df.columns
        if column != "Name" and column in averaged.columns
    ]
    if "Segment_Number" not in ordered_columns:
        ordered_columns.append("Segment_Number")

    return averaged[ordered_columns]


def dataframe_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Build a formatted workbook suitable for nontechnical users."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            export_df = dataframe.rename(columns=METRIC_LABELS)
            export_df.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

            worksheet = writer.book[sheet_name[:31]]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            header_fill = PatternFill(
                fill_type="solid",
                fgColor="3659E3",
            )
            header_font = Font(
                color="FFFFFF",
                bold=True,
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            worksheet.row_dimensions[1].height = 28

            for column_index, column_cells in enumerate(
                worksheet.iter_cols(),
                start=1,
            ):
                maximum = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = min(max(maximum + 3, 12), 42)

    return output.getvalue()


def dataframes_to_csv_zip(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(
        output,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        for name, dataframe in sheets.items():
            export_df = dataframe.rename(columns=METRIC_LABELS)
            archive.writestr(
                f"{name}.csv",
                export_df.to_csv(index=False).encode("utf-8"),
            )
    return output.getvalue()


def result_summary(dataframe: pd.DataFrame) -> dict[str, float | None]:
    """Create meaningful high-level values from whole-file results."""
    data = dataframe.copy()

    if "Name" in data.columns:
        totals_rows = data[
            data["Name"].astype(str).str.upper().eq("TOTALS")
        ]
        non_total_rows = data[
            ~data["Name"].astype(str).str.upper().eq("TOTALS")
        ]
    else:
        totals_rows = pd.DataFrame()
        non_total_rows = data

    summary: dict[str, float | None] = {
        "files": float(len(non_total_rows)),
        "total": None,
        "ue": None,
        "lf": None,
        "rf": None,
    }

    for key, column in [
        ("total", "Total_Counts"),
        ("ue", "UE_Counts"),
        ("lf", "LF_Counts"),
        ("rf", "RF_Counts"),
    ]:
        if column not in data.columns:
            continue

        if not totals_rows.empty:
            summary[key] = mean_from_metric(totals_rows.iloc[0][column])
        else:
            numeric = pd.to_numeric(
                non_total_rows[column],
                errors="coerce",
            )
            summary[key] = float(numeric.sum()) if numeric.notna().any() else None

    return summary


def chartable_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Convert display-formatted result fields into chartable mean values."""
    output = dataframe.copy()

    for column in output.columns:
        if column == "Name":
            continue
        output[column] = output[column].map(mean_from_metric)

    return output


def render_result_chart(
    dataframe: pd.DataFrame,
    result_name: str,
    key_prefix: str = "",
) -> None:
    """Render charts with an explicit Vega-Lite specification.

    The explicit long-form data and fixed field names avoid Altair schema
    errors that can occur when st.line_chart or st.bar_chart infer fields from
    pivoted DataFrames, especially with Altair 6.
    """
    available = [
        metric
        for metric in AVAILABLE_METRICS
        if metric in dataframe.columns
    ]

    if not available or "Name" not in dataframe.columns:
        st.info("No chart-compatible result columns are available.")
        return

    default_metric = (
        "Total_Counts"
        if "Total_Counts" in available
        else available[0]
    )

    selected_metric = st.selectbox(
        "Metric to visualize",
        options=available,
        index=available.index(default_metric),
        format_func=lambda metric: METRIC_LABELS.get(metric, metric),
        key=f"chart_metric_{key_prefix}_{result_name}_{analysis_cycle}",
        help=(
            "The chart uses the mean value when a table cell contains "
            "a mean and standard deviation."
        ),
    )

    chart_df = chartable_dataframe(dataframe)
    chart_df = chart_df[
        ~chart_df["Name"].astype(str).str.upper().eq("TOTALS")
    ].copy()
    chart_df = chart_df.dropna(subset=[selected_metric])

    if chart_df.empty:
        st.info("The selected metric has no chartable values.")
        return

    metric_title = METRIC_LABELS.get(selected_metric, selected_metric)

    try:
        if result_name == "Segment_Results":
            chart_df["Segment"] = pd.to_numeric(
                chart_df["Name"]
                .astype(str)
                .str.extract(r"Segment\s+(\d+)", expand=False),
                errors="coerce",
            )
            chart_df["Session"] = (
                chart_df["Name"]
                .astype(str)
                .str.replace(
                    r"\s*Segment\s+\d+\s*$",
                    "",
                    regex=True,
                )
                .str.strip()
            )

            # Averaged segment output is named only "Segment 1", etc.,
            # which leaves an empty session label after removing the suffix.
            chart_df.loc[
                chart_df["Session"].eq(""),
                "Session",
            ] = "Average"

            segment_chart = (
                chart_df[["Segment", "Session", selected_metric]]
                .dropna(subset=["Segment", selected_metric])
                .rename(columns={selected_metric: "Value"})
            )
            segment_chart["Segment"] = segment_chart["Segment"].astype(int)
            segment_chart["Session"] = segment_chart["Session"].astype(str)
            segment_chart["Value"] = pd.to_numeric(
                segment_chart["Value"],
                errors="coerce",
            )
            segment_chart = segment_chart.dropna(subset=["Value"])

            if not segment_chart.empty:
                st.vega_lite_chart(
                    segment_chart,
                    {
                        "mark": {
                            "type": "line",
                            "point": True,
                            "tooltip": True,
                        },
                        "encoding": {
                            "x": {
                                "field": "Segment",
                                "type": "quantitative",
                                "title": "Segment",
                                "axis": {"tickMinStep": 1},
                            },
                            "y": {
                                "field": "Value",
                                "type": "quantitative",
                                "title": metric_title,
                                "scale": {"zero": False},
                            },
                            "color": {
                                "field": "Session",
                                "type": "nominal",
                                "title": "Session",
                            },
                            "tooltip": [
                                {
                                    "field": "Session",
                                    "type": "nominal",
                                    "title": "Session",
                                },
                                {
                                    "field": "Segment",
                                    "type": "quantitative",
                                    "title": "Segment",
                                },
                                {
                                    "field": "Value",
                                    "type": "quantitative",
                                    "title": metric_title,
                                },
                            ],
                        },
                    },
                    width="stretch",
                    height=420,
                    key=f"segment_vega_{key_prefix}_{result_name}_{selected_metric}_{analysis_cycle}",
                )
                return

        whole_chart = (
            chart_df[["Name", selected_metric]]
            .rename(
                columns={
                    "Name": "Session",
                    selected_metric: "Value",
                }
            )
        )
        whole_chart["Session"] = whole_chart["Session"].astype(str)
        whole_chart["Value"] = pd.to_numeric(
            whole_chart["Value"],
            errors="coerce",
        )
        whole_chart = whole_chart.dropna(subset=["Value"])

        if whole_chart.empty:
            st.info("The selected metric has no chartable numeric values.")
            return

        st.vega_lite_chart(
            whole_chart,
            {
                "mark": {
                    "type": "bar",
                    "cornerRadiusTopLeft": 5,
                    "cornerRadiusTopRight": 5,
                    "tooltip": True,
                },
                "encoding": {
                    "x": {
                        "field": "Session",
                        "type": "nominal",
                        "title": "Session",
                        "sort": None,
                        "axis": {
                            "labelAngle": -35,
                            "labelLimit": 180,
                        },
                    },
                    "y": {
                        "field": "Value",
                        "type": "quantitative",
                        "title": metric_title,
                    },
                    "tooltip": [
                        {
                            "field": "Session",
                            "type": "nominal",
                            "title": "Session",
                        },
                        {
                            "field": "Value",
                            "type": "quantitative",
                            "title": metric_title,
                        },
                    ],
                },
            },
            width="stretch",
            height=420,
            key=f"whole_vega_{key_prefix}_{result_name}_{selected_metric}_{analysis_cycle}",
        )

    except Exception as chart_error:
        # A chart must never stop users from reaching their data or downloads.
        st.warning(
            "The visual chart could not be displayed, but the result tables "
            "and downloads remain available."
        )
        with st.expander("Chart diagnostic details"):
            st.code(str(chart_error))

def clear_analysis_state() -> None:
    for key in [
        "midipy_results",
        "midipy_valid_names",
        "midipy_skipped_files",
        "midipy_analysis_signature",
        "midipy_analysis_settings_signature",
        "midipy_last_settings",
    ]:
        st.session_state.pop(key, None)



def start_new_analysis() -> None:
    """Reset results, uploaded files, and all configurable controls."""
    clear_analysis_state()
    st.session_state["analysis_cycle"] = (
        int(st.session_state.get("analysis_cycle", 0)) + 1
    )
    st.rerun()






# =============================================================================
# COMPARISON-AWARE INTERACTION HELPERS
# =============================================================================

ROLE_THERAPIST = "Music therapist"
ROLE_PARTICIPANT = "Participant"
ROLE_SINGLE = "Performer"


def clear_analysis_state() -> None:
    for key in [
        "midipy_results",
        "midipy_valid_names",
        "midipy_skipped_files",
        "midipy_analysis_signature",
        "midipy_analysis_settings_signature",
        "midipy_last_settings",
    ]:
        st.session_state.pop(key, None)


def start_new_analysis() -> None:
    """Reset results, uploads, and configurable controls."""
    clear_analysis_state()
    st.session_state["analysis_cycle"] = (
        int(st.session_state.get("analysis_cycle", 0)) + 1
    )
    st.rerun()


def has_existing_work() -> bool:
    upload_keys = [
        f"midi_upload_{analysis_cycle}",
        f"therapist_upload_{analysis_cycle}",
        f"participant_upload_{analysis_cycle}",
    ]

    return any(
        bool(st.session_state.get(key))
        for key in upload_keys
    ) or bool(st.session_state.get("midipy_results"))


@st.dialog("Start a new analysis?")
def confirm_new_analysis() -> None:
    st.write(
        "This will remove the uploaded files, current settings, and displayed "
        "results from this browser session."
    )
    cancel_column, reset_column = st.columns(2)

    with cancel_column:
        st.caption("Close this window to keep the current analysis.")

    with reset_column:
        if st.button(
            "Clear and start new",
            type="primary",
            use_container_width=True,
            key=f"confirm_reset_{analysis_cycle}",
        ):
            start_new_analysis()


def render_progress_stepper(active_step: int) -> None:
    labels = ["Upload", "Configure", "Analyze", "Review"]
    parts: list[str] = [
        '<div class="mp-progress" aria-label="Analysis progress">'
    ]

    for index, label in enumerate(labels, start=1):
        if index < active_step:
            state_class = "is-complete"
            marker = "✓"
            state_text = "completed"
        elif index == active_step:
            state_class = "is-active"
            marker = str(index)
            state_text = "current"
        else:
            state_class = ""
            marker = str(index)
            state_text = "upcoming"

        parts.append(
            f'<div class="mp-progress-step {state_class}" '
            f'aria-label="{label}, {state_text}">'
            f"<span>{marker}</span><strong>{label}</strong></div>"
        )

        if index < len(labels):
            line_class = "is-complete" if index < active_step else ""
            parts.append(
                f'<div class="mp-progress-line {line_class}" '
                'aria-hidden="true"></div>'
            )

    parts.append("</div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


def format_file_status_table(
    precheck_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for row in precheck_rows:
        details = row["Details"]

        if row["Status"] == "Ready":
            status = "✓ Ready"
        elif "Empty" in details:
            status = "⚠ Empty file"
        elif "header" in details.lower():
            status = "✕ Invalid MIDI header"
        else:
            status = "⚠ Needs attention"

        rows.append(
            {
                "File": row["File"],
                "Size": row["Size"],
                "Status": status,
                "Details": details,
            }
        )

    return pd.DataFrame(rows)


def selected_custom_metrics(include_feet: bool) -> list[str]:
    """Show only measures relevant to the selected instrument."""
    selected: list[str] = []
    group_columns = st.columns(3)

    for column, (group_name, metrics) in zip(
        group_columns,
        METRIC_GROUPS.items(),
    ):
        visible_metrics = [
            metric
            for metric in metrics
            if include_feet or metric not in FOOT_METRICS
        ]

        with column:
            st.markdown(f"**{group_name}**")

            for metric in visible_metrics:
                checked = st.checkbox(
                    METRIC_LABELS[metric],
                    value=True,
                    help=METRIC_HELP.get(metric),
                    key=f"metric_{metric}_{analysis_cycle}",
                )

                if checked:
                    selected.append(metric)

    return selected


def combined_upload_signature(
    uploads_by_role: dict[str, list[Any]],
) -> str:
    """Create one stable signature for single or comparison uploads."""
    role_signatures = {
        role: upload_signature(files)
        for role, files in sorted(uploads_by_role.items())
    }
    return settings_signature(role_signatures)


def run_dataset_analysis(
    *,
    uploaded_files: list[Any],
    temporary_path: Path,
    dataset_slug: str,
    metrics_argument: list[str],
    run_whole: bool,
    run_segments: bool,
    number_of_segments: int,
    average_segments: bool,
    ue_keys: list[int],
    left_foot_key: int,
    right_foot_key: int,
) -> tuple[
    dict[str, pd.DataFrame],
    list[str],
    list[tuple[str, str]],
]:
    """Validate and analyze one performer dataset."""
    midi_folder = temporary_path / f"{dataset_slug}_validated_midi"
    midi_folder.mkdir(parents=True, exist_ok=True)

    valid_names, skipped_files = validate_and_save_uploads(
        uploaded_files,
        midi_folder,
    )

    if not valid_names:
        return {}, valid_names, skipped_files

    results: dict[str, pd.DataFrame] = {}

    if run_whole:
        whole_df = parser(
            source=str(midi_folder),
            metrics=metrics_argument,
            output_format="csv",
            save_path=str(
                temporary_path / f"{dataset_slug}_whole_results"
            ),
            ue_keys=ue_keys,
            lf_key=left_foot_key,
            rf_key=right_foot_key,
        )
        results["Whole_File_Results"] = whole_df

    if run_segments:
        segment_df = parser_segments(
            source=str(midi_folder),
            metrics=metrics_argument,
            output_format="csv",
            save_path=str(
                temporary_path / f"{dataset_slug}_segment_results"
            ),
            num_segments=number_of_segments,
            mean_segments=False,
            ue_keys=ue_keys,
            lf_key=left_foot_key,
            rf_key=right_foot_key,
        )

        if average_segments:
            segment_df = average_segment_rows(segment_df)

        results["Segment_Results"] = segment_df

    return results, valid_names, skipped_files


def aggregate_metric_value(
    dataframe: pd.DataFrame,
    metric: str,
) -> float | None:
    """Aggregate one measure across all files without implying quality."""
    if metric not in dataframe.columns:
        return None

    working = dataframe.copy()

    if "Name" in working.columns:
        totals = working[
            working["Name"].astype(str).str.upper().eq("TOTALS")
        ]
        observations = working[
            ~working["Name"].astype(str).str.upper().eq("TOTALS")
        ]
    else:
        totals = pd.DataFrame()
        observations = working

    if metric in COUNT_METRICS:
        if not totals.empty:
            return mean_from_metric(totals.iloc[0][metric])

        values = observations[metric].map(mean_from_metric).dropna()
        return float(values.sum()) if not values.empty else None

    values = observations[metric].map(mean_from_metric).dropna()
    return float(values.mean()) if not values.empty else None


def primary_result_dataframe(
    dataset_results: dict[str, pd.DataFrame],
) -> pd.DataFrame | None:
    """Prefer whole-file results; otherwise use segment results."""
    if "Whole_File_Results" in dataset_results:
        return dataset_results["Whole_File_Results"]

    return dataset_results.get("Segment_Results")


def build_comparison_summary(
    therapist_results: dict[str, pd.DataFrame],
    participant_results: dict[str, pd.DataFrame],
    selected_metrics: list[str],
) -> pd.DataFrame:
    therapist_df = primary_result_dataframe(therapist_results)
    participant_df = primary_result_dataframe(participant_results)

    if therapist_df is None or participant_df is None:
        return pd.DataFrame()

    rows: list[dict[str, Any]] = []

    for metric in selected_metrics:
        therapist_value = aggregate_metric_value(
            therapist_df,
            metric,
        )
        participant_value = aggregate_metric_value(
            participant_df,
            metric,
        )

        if therapist_value is None or participant_value is None:
            continue

        difference = participant_value - therapist_value
        relative_difference = (
            difference / abs(therapist_value) * 100
            if therapist_value != 0
            else None
        )

        rows.append(
            {
                "Metric key": metric,
                "Measure": METRIC_LABELS.get(metric, metric),
                ROLE_THERAPIST: therapist_value,
                ROLE_PARTICIPANT: participant_value,
                "Difference": difference,
                "Relative difference (%)": relative_difference,
            }
        )

    return pd.DataFrame(rows)


def comparison_export_tables(
    datasets: dict[str, dict[str, pd.DataFrame]],
    comparison_summary: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    export_tables: dict[str, pd.DataFrame] = {}

    role_prefixes = {
        ROLE_THERAPIST: "Therapist",
        ROLE_PARTICIPANT: "Participant",
        ROLE_SINGLE: "Performer",
    }

    for role, dataset_results in datasets.items():
        prefix = role_prefixes.get(role, role)

        for result_name, dataframe in dataset_results.items():
            result_suffix = (
                "Whole"
                if result_name == "Whole_File_Results"
                else "Segments"
            )
            export_tables[f"{prefix} {result_suffix}"] = dataframe

    if not comparison_summary.empty:
        export_tables["Comparison Summary"] = (
            comparison_summary
            .drop(columns=["Metric key"], errors="ignore")
        )

    return export_tables


def render_comparison_chart(
    comparison_summary: pd.DataFrame,
) -> None:
    if comparison_summary.empty:
        st.info("No shared measures are available for comparison.")
        return

    metric_options = comparison_summary["Metric key"].tolist()

    selected_metric = st.selectbox(
        "Measure to compare",
        options=metric_options,
        format_func=lambda metric: METRIC_LABELS.get(metric, metric),
        key=f"comparison_metric_{analysis_cycle}",
    )

    selected_row = comparison_summary[
        comparison_summary["Metric key"].eq(selected_metric)
    ].iloc[0]

    chart_data = pd.DataFrame(
        {
            "Performer": [
                ROLE_THERAPIST,
                ROLE_PARTICIPANT,
            ],
            "Value": [
                selected_row[ROLE_THERAPIST],
                selected_row[ROLE_PARTICIPANT],
            ],
        }
    )

    metric_title = METRIC_LABELS.get(
        selected_metric,
        selected_metric,
    )

    st.vega_lite_chart(
        chart_data,
        {
            "mark": {
                "type": "bar",
                "cornerRadiusTopLeft": 7,
                "cornerRadiusTopRight": 7,
                "tooltip": True,
            },
            "encoding": {
                "x": {
                    "field": "Performer",
                    "type": "nominal",
                    "title": None,
                    "sort": [
                        ROLE_THERAPIST,
                        ROLE_PARTICIPANT,
                    ],
                },
                "y": {
                    "field": "Value",
                    "type": "quantitative",
                    "title": metric_title,
                },
                "color": {
                    "field": "Performer",
                    "type": "nominal",
                    "legend": None,
                    "scale": {
                        "domain": [
                            ROLE_THERAPIST,
                            ROLE_PARTICIPANT,
                        ],
                        "range": [
                            "#3157D5",
                            "#1A8B7E",
                        ],
                    },
                },
                "tooltip": [
                    {
                        "field": "Performer",
                        "type": "nominal",
                        "title": "Performer",
                    },
                    {
                        "field": "Value",
                        "type": "quantitative",
                        "title": metric_title,
                        "format": ".2f",
                    },
                ],
            },
        },
        width="stretch",
        height=350,
        key=f"comparison_bar_{selected_metric}_{analysis_cycle}",
    )


def segment_comparison_dataframe(
    therapist_segments: pd.DataFrame,
    participant_segments: pd.DataFrame,
    metric: str,
) -> pd.DataFrame:
    rows: list[pd.DataFrame] = []

    for role, dataframe in [
        (ROLE_THERAPIST, therapist_segments),
        (ROLE_PARTICIPANT, participant_segments),
    ]:
        if "Name" not in dataframe.columns or metric not in dataframe.columns:
            continue

        working = dataframe[["Name", metric]].copy()
        working["Segment"] = pd.to_numeric(
            working["Name"]
            .astype(str)
            .str.extract(r"Segment\s+(\d+)", expand=False),
            errors="coerce",
        )
        working["Value"] = working[metric].map(mean_from_metric)
        working = working.dropna(subset=["Segment", "Value"])

        if working.empty:
            continue

        averaged = (
            working.groupby("Segment", as_index=False)["Value"]
            .mean()
            .sort_values("Segment")
        )
        averaged["Performer"] = role
        rows.append(averaged)

    if not rows:
        return pd.DataFrame()

    return pd.concat(rows, ignore_index=True)


def render_segment_comparison(
    therapist_results: dict[str, pd.DataFrame],
    participant_results: dict[str, pd.DataFrame],
    selected_metrics: list[str],
) -> None:
    therapist_segments = therapist_results.get("Segment_Results")
    participant_segments = participant_results.get("Segment_Results")

    if therapist_segments is None or participant_segments is None:
        return

    shared_metrics = [
        metric
        for metric in selected_metrics
        if (
            metric in therapist_segments.columns
            and metric in participant_segments.columns
        )
    ]

    if not shared_metrics:
        return

    st.subheader("Segment trajectories")

    selected_metric = st.selectbox(
        "Measure across segments",
        options=shared_metrics,
        format_func=lambda metric: METRIC_LABELS.get(metric, metric),
        key=f"comparison_segment_metric_{analysis_cycle}",
    )

    chart_data = segment_comparison_dataframe(
        therapist_segments,
        participant_segments,
        selected_metric,
    )

    if chart_data.empty:
        st.info("No segment values are available for this measure.")
        return

    metric_title = METRIC_LABELS.get(
        selected_metric,
        selected_metric,
    )

    st.vega_lite_chart(
        chart_data,
        {
            "mark": {
                "type": "line",
                "point": True,
                "tooltip": True,
                "strokeWidth": 3,
            },
            "encoding": {
                "x": {
                    "field": "Segment",
                    "type": "quantitative",
                    "title": "Segment",
                    "axis": {"tickMinStep": 1},
                },
                "y": {
                    "field": "Value",
                    "type": "quantitative",
                    "title": metric_title,
                    "scale": {"zero": False},
                },
                "color": {
                    "field": "Performer",
                    "type": "nominal",
                    "title": None,
                    "scale": {
                        "domain": [
                            ROLE_THERAPIST,
                            ROLE_PARTICIPANT,
                        ],
                        "range": [
                            "#3157D5",
                            "#1A8B7E",
                        ],
                    },
                },
                "tooltip": [
                    {
                        "field": "Performer",
                        "type": "nominal",
                        "title": "Performer",
                    },
                    {
                        "field": "Segment",
                        "type": "quantitative",
                        "title": "Segment",
                    },
                    {
                        "field": "Value",
                        "type": "quantitative",
                        "title": metric_title,
                        "format": ".2f",
                    },
                ],
            },
        },
        width="stretch",
        height=380,
        key=f"comparison_segment_chart_{selected_metric}_{analysis_cycle}",
    )


def render_dataset_summary(
    dataset_results: dict[str, pd.DataFrame],
    valid_names: list[str],
    skipped_files: list[tuple[str, str]],
    include_feet: bool,
) -> None:
    whole_dataframe = dataset_results.get("Whole_File_Results")

    if whole_dataframe is not None:
        summary = result_summary(whole_dataframe)

        if include_feet:
            columns = st.columns(5)
            columns[0].metric(
                "Files analyzed",
                int(summary["files"] or len(valid_names)),
            )
            columns[1].metric(
                "Total notes",
                f"{int(summary['total']):,}"
                if summary["total"] is not None
                else "—",
            )
            columns[2].metric(
                "UE notes",
                f"{int(summary['ue']):,}"
                if summary["ue"] is not None
                else "—",
            )
            columns[3].metric(
                "LF notes",
                f"{int(summary['lf']):,}"
                if summary["lf"] is not None
                else "—",
            )
            columns[4].metric(
                "RF notes",
                f"{int(summary['rf']):,}"
                if summary["rf"] is not None
                else "—",
            )
        else:
            columns = st.columns(3)
            columns[0].metric(
                "Files analyzed",
                int(summary["files"] or len(valid_names)),
            )
            columns[1].metric(
                "Total notes",
                f"{int(summary['total']):,}"
                if summary["total"] is not None
                else "—",
            )
            columns[2].metric(
                "UE notes",
                f"{int(summary['ue']):,}"
                if summary["ue"] is not None
                else "—",
            )
    else:
        columns = st.columns(3)
        columns[0].metric("Files analyzed", len(valid_names))
        columns[1].metric(
            "Segment rows",
            len(dataset_results.get("Segment_Results", [])),
        )
        columns[2].metric("Files skipped", len(skipped_files))


def render_dataset_details(
    *,
    role: str,
    dataset_results: dict[str, pd.DataFrame],
    valid_names: list[str],
    skipped_files: list[tuple[str, str]],
    include_feet: bool,
) -> None:
    role_key = re.sub(r"[^a-z0-9]+", "_", role.lower()).strip("_")

    render_dataset_summary(
        dataset_results,
        valid_names,
        skipped_files,
        include_feet,
    )

    st.subheader("Visual overview")

    result_options = list(dataset_results.keys())
    selected_result_name = st.radio(
        "Result set",
        options=result_options,
        horizontal=True,
        format_func=lambda name: (
            "Whole-file results"
            if name == "Whole_File_Results"
            else "Segment results"
        ),
        key=f"result_set_{role_key}_{analysis_cycle}",
    )

    render_result_chart(
        dataset_results[selected_result_name],
        selected_result_name,
        key_prefix=role_key,
    )

    st.subheader("Detailed data")

    for result_name, dataframe in dataset_results.items():
        label = (
            "Whole-file results"
            if result_name == "Whole_File_Results"
            else "Segment results"
        )

        with st.expander(label, expanded=len(dataset_results) == 1):
            display_dataframe(dataframe)
            st.caption(
                f"{len(dataframe):,} row(s) × "
                f"{len(dataframe.columns):,} column(s)"
            )

    with st.expander("File quality"):
        quality_left, quality_right = st.columns(2)

        with quality_left:
            st.markdown("**Successfully analyzed**")
            for filename in valid_names:
                st.write(f"✓ {filename}")

        with quality_right:
            st.markdown("**Skipped during validation**")
            if skipped_files:
                st.dataframe(
                    pd.DataFrame(
                        skipped_files,
                        columns=["File", "Reason"],
                    ),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.write("✓ No files were skipped.")


# =============================================================================
# COMPACT HEADER
# =============================================================================

header_content, header_actions = st.columns(
    [5.2, 1.8],
    vertical_alignment="center",
)

with header_content:
    st.markdown(
        """
        <h1 class="mp-header-title">MidiPy Analysis Studio</h1>
        <p class="mp-header-subtitle">
            Validate MIDI files, configure mappings, compare performers,
            and export clear results.
        </p>
        """,
        unsafe_allow_html=True,
    )

with header_actions:
    help_column, reset_column = st.columns([0.8, 1.25])

    with help_column:
        with st.popover("Help", use_container_width=True):
            st.markdown(
                """
                **Workflow**

                1. Choose a single analysis or comparison.
                2. Upload the required MIDI files.
                3. Confirm instrument and note mappings.
                4. Choose whole-file or segment analysis.
                5. Review and export the results.

                **Terminology**

                - **UE:** upper extremity
                - **LF:** left foot
                - **RF:** right foot
                - **Velocity:** MIDI performance intensity
                - **Asynchrony:** timing difference from the quantized beat
                """
            )

    with reset_column:
        if st.button(
            "New analysis",
            use_container_width=True,
            key=f"new_analysis_{analysis_cycle}",
        ):
            if has_existing_work():
                confirm_new_analysis()
            else:
                start_new_analysis()

progress_placeholder = st.empty()


# =============================================================================
# ANALYSIS SETUP
# =============================================================================

st.markdown(
    """
    <h2 class="mp-section-title">Analysis setup</h2>
    <p class="mp-section-copy">
        Choose whether to analyze one performer or compare two performer datasets.
    </p>
    """,
    unsafe_allow_html=True,
)

with st.container(border=True):
    analysis_design = st.radio(
        "Would you like to compare a music therapist and a participant?",
        options=[
            "Single-performer analysis",
            "Therapist–participant comparison",
        ],
        index=0,
        horizontal=True,
        key=f"analysis_design_{analysis_cycle}",
    )

    comparison_enabled = (
        analysis_design == "Therapist–participant comparison"
    )

    if comparison_enabled:
        st.markdown(
            """
            <div class="mp-analysis-mode-note">
                The same instrument mapping, segment settings, and result measures
                will be applied to both datasets. This keeps the comparison
                consistent and easier to interpret.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="mp-analysis-mode-note">
                Analyze one performer or one collection of sessions using the
                standard MidiPy workflow.
            </div>
            """,
            unsafe_allow_html=True,
        )


# =============================================================================
# 1. UPLOAD
# =============================================================================

st.markdown(
    """
    <h2 class="mp-section-title">1. Upload MIDI files</h2>
    <p class="mp-section-copy">
        Add all sessions that should be included in this analysis.
    </p>
    """,
    unsafe_allow_html=True,
)

uploads_by_role: dict[str, list[Any]] = {}
prechecks_by_role: dict[str, list[dict[str, Any]]] = {}
ready_by_role: dict[str, int] = {}
attention_by_role: dict[str, int] = {}
size_by_role: dict[str, int] = {}

if comparison_enabled:
    therapist_column, participant_column = st.columns(2, gap="large")

    with therapist_column:
        with st.container(border=True):
            st.markdown(
                """
                <div class="mp-role-heading">
                    <span class="mp-role-badge therapist">MT</span>
                    <span class="mp-role-title">Music therapist</span>
                </div>
                <p class="mp-role-copy">
                    Upload the therapist's MIDI sessions.
                </p>
                """,
                unsafe_allow_html=True,
            )

            therapist_files = st.file_uploader(
                "Music therapist MIDI files",
                type=["mid", "midi"],
                accept_multiple_files=True,
                help=(
                    "Supported formats: .mid and .midi. "
                    "Maximum 200 MB per file."
                ),
                key=f"therapist_upload_{analysis_cycle}",
                label_visibility="collapsed",
            )

            (
                therapist_precheck,
                therapist_ready,
                therapist_size,
            ) = precheck_uploads(therapist_files)

            therapist_attention = (
                len(therapist_precheck) - therapist_ready
            )

            uploads_by_role[ROLE_THERAPIST] = therapist_files
            prechecks_by_role[ROLE_THERAPIST] = therapist_precheck
            ready_by_role[ROLE_THERAPIST] = therapist_ready
            attention_by_role[ROLE_THERAPIST] = therapist_attention
            size_by_role[ROLE_THERAPIST] = therapist_size

            if therapist_files:
                st.markdown(
                    f"""
                    <div class="mp-upload-summary" role="status">
                        <span class="mp-ready">
                            <strong>✓ {therapist_ready}</strong> ready
                        </span>
                        <span class="mp-attention">
                            <strong>⚠ {therapist_attention}</strong> need attention
                        </span>
                        <span>
                            <strong>{human_file_size(therapist_size)}</strong>
                        </span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                with st.expander(
                    "View therapist file details",
                    expanded=therapist_attention > 0,
                ):
                    st.dataframe(
                        format_file_status_table(therapist_precheck),
                        use_container_width=True,
                        hide_index=True,
                    )
            else:
                st.caption("No therapist files selected.")

    with participant_column:
        with st.container(border=True):
            st.markdown(
                """
                <div class="mp-role-heading">
                    <span class="mp-role-badge participant">P</span>
                    <span class="mp-role-title">Participant</span>
                </div>
                <p class="mp-role-copy">
                    Upload the participant's MIDI sessions.
                </p>
                """,
                unsafe_allow_html=True,
            )

            participant_files = st.file_uploader(
                "Participant MIDI files",
                type=["mid", "midi"],
                accept_multiple_files=True,
                help=(
                    "Supported formats: .mid and .midi. "
                    "Maximum 200 MB per file."
                ),
                key=f"participant_upload_{analysis_cycle}",
                label_visibility="collapsed",
            )

            (
                participant_precheck,
                participant_ready,
                participant_size,
            ) = precheck_uploads(participant_files)

            participant_attention = (
                len(participant_precheck) - participant_ready
            )

            uploads_by_role[ROLE_PARTICIPANT] = participant_files
            prechecks_by_role[ROLE_PARTICIPANT] = participant_precheck
            ready_by_role[ROLE_PARTICIPANT] = participant_ready
            attention_by_role[ROLE_PARTICIPANT] = participant_attention
            size_by_role[ROLE_PARTICIPANT] = participant_size

            if participant_files:
                st.markdown(
                    f"""
                    <div class="mp-upload-summary" role="status">
                        <span class="mp-ready">
                            <strong>✓ {participant_ready}</strong> ready
                        </span>
                        <span class="mp-attention">
                            <strong>⚠ {participant_attention}</strong> need attention
                        </span>
                        <span>
                            <strong>{human_file_size(participant_size)}</strong>
                        </span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                with st.expander(
                    "View participant file details",
                    expanded=participant_attention > 0,
                ):
                    st.dataframe(
                        format_file_status_table(participant_precheck),
                        use_container_width=True,
                        hide_index=True,
                    )
            else:
                st.caption("No participant files selected.")

    st.markdown(
        """
        <div class="mp-comparison-callout">
            Upload at least one valid MIDI file for each performer. The two
            datasets may contain different numbers of sessions; group-level
            values will be calculated consistently for comparison.
        </div>
        """,
        unsafe_allow_html=True,
    )

else:
    with st.container(border=True):
        performer_files = st.file_uploader(
            "Drag and drop MIDI files or browse",
            type=["mid", "midi"],
            accept_multiple_files=True,
            help=(
                "Supported formats: .mid and .midi. "
                "Maximum 200 MB per file."
            ),
            key=f"midi_upload_{analysis_cycle}",
        )

        (
            performer_precheck,
            performer_ready,
            performer_size,
        ) = precheck_uploads(performer_files)

        performer_attention = (
            len(performer_precheck) - performer_ready
        )

        uploads_by_role[ROLE_SINGLE] = performer_files
        prechecks_by_role[ROLE_SINGLE] = performer_precheck
        ready_by_role[ROLE_SINGLE] = performer_ready
        attention_by_role[ROLE_SINGLE] = performer_attention
        size_by_role[ROLE_SINGLE] = performer_size

        if performer_files:
            st.markdown(
                f"""
                <div class="mp-upload-summary" role="status">
                    <span class="mp-ready">
                        <strong>✓ {performer_ready}</strong> file(s) ready
                    </span>
                    <span class="mp-attention">
                        <strong>⚠ {performer_attention}</strong> need attention
                    </span>
                    <span>
                        <strong>{human_file_size(performer_size)}</strong> total
                    </span>
                </div>
                """,
                unsafe_allow_html=True,
            )

            with st.expander(
                "View file details",
                expanded=performer_attention > 0,
            ):
                st.dataframe(
                    format_file_status_table(performer_precheck),
                    use_container_width=True,
                    hide_index=True,
                )
        else:
            st.caption(
                "No files selected · Supported: MID, MIDI · "
                "Maximum 200 MB each"
            )

st.caption(
    "Privacy: use de-identified research files and follow your institution's "
    "data-handling requirements."
)

uploads_ready = (
    all(ready_by_role.get(role, 0) > 0 for role in uploads_by_role)
    if uploads_by_role
    else False
)

if not uploads_ready:
    with progress_placeholder:
        render_progress_stepper(active_step=1)

    if comparison_enabled:
        st.info(
            "Add at least one valid MIDI file for both the music therapist "
            "and the participant to continue."
        )

    st.markdown(
        """
        <div class="mp-footer">
            MidiPy Analysis Studio · Human-centred MIDI analysis
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()


# =============================================================================
# 2. CONFIGURE
# =============================================================================

st.markdown(
    """
    <h2 class="mp-section-title">2. Configure the analysis</h2>
    <p class="mp-section-copy">
        Defaults are prepared. Change only what is required for this dataset.
    </p>
    """,
    unsafe_allow_html=True,
)

if comparison_enabled:
    st.markdown(
        """
        <div class="mp-comparison-callout">
            These settings apply to both the music therapist and participant.
            If the performers used different instruments or incompatible note
            mappings, analyze them separately rather than comparing them directly.
        </div>
        """,
        unsafe_allow_html=True,
    )

mapping_column, scope_column = st.columns(
    [1.15, 1],
    gap="large",
)

with mapping_column:
    with st.container(border=True):
        st.markdown(
            """
            <h3 class="mp-card-heading">Instrument and note mapping</h3>
            <p class="mp-card-copy">
                Foot mappings are used for drum-set analysis and omitted for
                guitar or other instruments.
            </p>
            """,
            unsafe_allow_html=True,
        )

        instrument_mode = st.radio(
            "Instrument setup",
            options=[
                "Drum set",
                "Guitar or other instrument",
            ],
            index=0,
            key=f"instrument_mode_{analysis_cycle}",
        )

        include_feet = instrument_mode == "Drum set"

        ue_keys = st.multiselect(
            "Upper-extremity notes",
            options=list(range(128)),
            default=DEFAULT_UE_KEYS,
            placeholder="Search or add a MIDI note number",
            help=(
                "Choose all note numbers assigned to the upper extremities."
            ),
            key=f"ue_note_values_{analysis_cycle}",
        )
        st.caption(f"{len(ue_keys)} upper-extremity note(s) selected")

        if include_feet:
            foot_left, foot_right = st.columns(2)

            with foot_left:
                left_foot_key = int(
                    st.number_input(
                        "Left foot",
                        min_value=0,
                        max_value=127,
                        value=44,
                        step=1,
                        key=f"left_foot_key_{analysis_cycle}",
                    )
                )

            with foot_right:
                right_foot_key = int(
                    st.number_input(
                        "Right foot",
                        min_value=0,
                        max_value=127,
                        value=36,
                        step=1,
                        key=f"right_foot_key_{analysis_cycle}",
                    )
                )

            mapping_conflicts: list[str] = []

            if left_foot_key in ue_keys:
                mapping_conflicts.append(
                    f"Note {left_foot_key} is assigned to both UE and left foot."
                )

            if right_foot_key in ue_keys:
                mapping_conflicts.append(
                    f"Note {right_foot_key} is assigned to both UE and right foot."
                )

            if left_foot_key == right_foot_key:
                mapping_conflicts.append(
                    f"Note {left_foot_key} is assigned to both feet."
                )

            if mapping_conflicts:
                st.markdown(
                    '<div class="mp-inline-warning">⚠ '
                    + "<br>⚠ ".join(mapping_conflicts)
                    + "</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div class="mp-inline-ok">'
                    '✓ No mapping conflicts detected.</div>',
                    unsafe_allow_html=True,
                )

        else:
            left_foot_key = 44
            right_foot_key = 36
            st.info(
                "Foot mappings and LF/RF measures are disabled for this "
                "instrument."
            )

with scope_column:
    with st.container(border=True):
        st.markdown(
            """
            <h3 class="mp-card-heading">Analysis scope</h3>
            <p class="mp-card-copy">
                Choose the level of detail needed for this analysis.
            </p>
            """,
            unsafe_allow_html=True,
        )

        analysis_mode = st.radio(
            "Choose what to analyze",
            options=[
                "Whole files + segments",
                "Whole files only",
                "Segments only",
            ],
            index=0,
            key=f"analysis_mode_{analysis_cycle}",
            label_visibility="collapsed",
        )

        run_whole = analysis_mode in {
            "Whole files + segments",
            "Whole files only",
        }
        run_segments = analysis_mode in {
            "Whole files + segments",
            "Segments only",
        }

        if analysis_mode == "Whole files + segments":
            st.caption(
                "Recommended · Provides session summaries and change over time."
            )
        elif analysis_mode == "Whole files only":
            st.caption(
                "Best for one summary row per complete MIDI session."
            )
        else:
            st.caption(
                "Best for examining changes across sections of each session."
            )

        if run_segments:
            with st.container(border=True):
                st.markdown("**Segment settings**")
                number_of_segments = int(
                    st.slider(
                        "Number of segments",
                        min_value=2,
                        max_value=20,
                        value=5,
                        key=f"segment_count_{analysis_cycle}",
                    )
                )
                average_segments = st.checkbox(
                    "Average matching segments across files",
                    value=False,
                    help=(
                        "Combines all Segment 1 rows, all Segment 2 rows, "
                        "and so on."
                    ),
                    key=f"average_segments_{analysis_cycle}",
                )
        else:
            number_of_segments = 5
            average_segments = False

with st.container(border=True):
    st.markdown(
        """
        <h3 class="mp-card-heading">Results to include</h3>
        <p class="mp-card-copy">
            Use a prepared report or choose measures by category.
        </p>
        """,
        unsafe_allow_html=True,
    )

    result_preset = st.radio(
        "Report detail",
        options=[
            "Complete report",
            "Counts only",
            "Custom selection",
        ],
        index=0,
        key=f"metric_preset_{analysis_cycle}",
        label_visibility="collapsed",
    )

    available_for_instrument = (
        AVAILABLE_METRICS
        if include_feet
        else NON_FOOT_METRICS
    )

    if result_preset == "Complete report":
        selected_metrics = available_for_instrument.copy()

        if include_feet:
            st.caption(
                "12 measures selected · counts, velocity, and asynchrony"
            )
        else:
            st.caption(
                "6 non-foot measures selected · counts, velocity, "
                "and asynchrony"
            )

    elif result_preset == "Counts only":
        selected_metrics = (
            COUNT_METRICS.copy()
            if include_feet
            else NON_FOOT_COUNT_METRICS.copy()
        )

        if include_feet:
            st.caption(
                "4 measures selected · total, UE, LF, and RF note counts"
            )
        else:
            st.caption(
                "2 non-foot measures selected · total and UE note counts"
            )

    else:
        st.divider()
        selected_metrics = selected_custom_metrics(
            include_feet=include_feet
        )
        st.caption(
            f"{len(selected_metrics)} of "
            f"{len(available_for_instrument)} available measures selected"
        )


# =============================================================================
# SETTINGS AND STICKY PRIMARY ACTION
# =============================================================================

current_settings = {
    "analysis_design": analysis_design,
    "instrument_mode": instrument_mode,
    "include_feet": include_feet,
    "ue_keys": ue_keys,
    "left_foot_key": left_foot_key if include_feet else None,
    "right_foot_key": right_foot_key if include_feet else None,
    "analysis_mode": analysis_mode,
    "number_of_segments": number_of_segments,
    "average_segments": average_segments,
    "selected_metrics": selected_metrics,
}

current_upload_signature = combined_upload_signature(
    uploads_by_role
)
current_settings_signature = settings_signature(
    current_settings
)

existing_results = st.session_state.get("midipy_results")
existing_upload_signature = st.session_state.get(
    "midipy_analysis_signature"
)
existing_settings_signature = st.session_state.get(
    "midipy_analysis_settings_signature"
)

existing_results_are_current = (
    bool(existing_results)
    and existing_upload_signature == current_upload_signature
    and existing_settings_signature == current_settings_signature
)

with st.container(border=True):
    st.markdown(
        '<div class="mp-sticky-marker"></div>',
        unsafe_allow_html=True,
    )

    action_information, action_button = st.columns(
        [4, 1.45],
        vertical_alignment="center",
    )

    with action_information:
        if comparison_enabled:
            status_text = (
                f"{ready_by_role[ROLE_THERAPIST]} therapist file(s) ready · "
                f"{ready_by_role[ROLE_PARTICIPANT]} participant file(s) ready · "
                f"{len(selected_metrics)} measure(s)"
            )
        else:
            status_text = (
                f"{ready_by_role[ROLE_SINGLE]} file(s) ready · "
                f"{len(selected_metrics)} measure(s)"
            )

        st.markdown(f"**{status_text}**")
        st.caption(
            "Complete MIDI validation runs when analysis begins."
        )

    with action_button:
        submitted = st.button(
            (
                "Run comparison"
                if comparison_enabled
                else "Analyze MIDI files"
            ),
            type="primary",
            use_container_width=True,
            disabled=(
                not uploads_ready
                or not ue_keys
                or not selected_metrics
            ),
            key=f"analyze_midi_files_{analysis_cycle}",
        )

if existing_results and not existing_results_are_current:
    st.warning(
        "Files or settings changed after the last analysis. "
        "Run the analysis again to refresh the results."
    )


# =============================================================================
# 3. ANALYZE
# =============================================================================

if submitted:
    with progress_placeholder:
        render_progress_stepper(active_step=3)

    with st.status(
        (
            "Preparing the comparison…"
            if comparison_enabled
            else "Preparing the analysis…"
        ),
        expanded=True,
    ) as analysis_status:
        try:
            with tempfile.TemporaryDirectory(
                prefix="midipy_dashboard_"
            ) as temporary:
                temporary_path = Path(temporary)

                metrics_argument = (
                    ["all"]
                    if (
                        include_feet
                        and set(selected_metrics) == set(AVAILABLE_METRICS)
                    )
                    else ["Name", *selected_metrics]
                )

                datasets: dict[
                    str,
                    dict[str, pd.DataFrame],
                ] = {}
                valid_names_by_role: dict[str, list[str]] = {}
                skipped_files_by_role: dict[
                    str,
                    list[tuple[str, str]],
                ] = {}

                for role, uploaded_files in uploads_by_role.items():
                    role_slug = re.sub(
                        r"[^a-z0-9]+",
                        "_",
                        role.lower(),
                    ).strip("_")

                    analysis_status.write(
                        f"Validating and analyzing {role.lower()} files."
                    )

                    (
                        dataset_results,
                        valid_names,
                        skipped_files,
                    ) = run_dataset_analysis(
                        uploaded_files=uploaded_files,
                        temporary_path=temporary_path,
                        dataset_slug=role_slug,
                        metrics_argument=metrics_argument,
                        run_whole=run_whole,
                        run_segments=run_segments,
                        number_of_segments=number_of_segments,
                        average_segments=average_segments,
                        ue_keys=ue_keys,
                        left_foot_key=left_foot_key,
                        right_foot_key=right_foot_key,
                    )

                    if not valid_names:
                        clear_analysis_state()
                        analysis_status.update(
                            label=(
                                f"No usable {role.lower()} MIDI files "
                                "were found."
                            ),
                            state="error",
                            expanded=True,
                        )
                        st.error(
                            f"None of the {role.lower()} files passed "
                            "complete MIDI validation."
                        )

                        if skipped_files:
                            st.dataframe(
                                pd.DataFrame(
                                    skipped_files,
                                    columns=["File", "Reason"],
                                ),
                                use_container_width=True,
                                hide_index=True,
                            )

                        st.stop()

                    datasets[role] = dataset_results
                    valid_names_by_role[role] = valid_names
                    skipped_files_by_role[role] = skipped_files

                analysis_payload = {
                    "mode": (
                        "comparison"
                        if comparison_enabled
                        else "single"
                    ),
                    "datasets": datasets,
                }

                st.session_state["midipy_results"] = analysis_payload
                st.session_state[
                    "midipy_valid_names"
                ] = valid_names_by_role
                st.session_state[
                    "midipy_skipped_files"
                ] = skipped_files_by_role
                st.session_state[
                    "midipy_analysis_signature"
                ] = current_upload_signature
                st.session_state[
                    "midipy_analysis_settings_signature"
                ] = current_settings_signature
                st.session_state[
                    "midipy_last_settings"
                ] = current_settings

                total_processed = sum(
                    len(names)
                    for names in valid_names_by_role.values()
                )

                analysis_status.update(
                    label=(
                        f"Analysis complete — "
                        f"{total_processed} file(s) processed."
                    ),
                    state="complete",
                    expanded=False,
                )

        except Exception as error:
            clear_analysis_state()
            analysis_status.update(
                label="The analysis could not be completed.",
                state="error",
                expanded=True,
            )
            st.error(
                "MidiPy encountered a problem while processing the files. "
                "Review the guidance below and try again."
            )
            st.markdown(
                """
                - Confirm that the files are genuine Standard MIDI files.
                - Remove empty or damaged files.
                - Confirm that mappings use values from 0 to 127.
                """
            )

            with st.expander("Technical details for support"):
                st.code(str(error))


# =============================================================================
# 4. REVIEW AND EXPORT
# =============================================================================

analysis_payload = st.session_state.get("midipy_results")
stored_upload_signature = st.session_state.get(
    "midipy_analysis_signature"
)
stored_settings_signature = st.session_state.get(
    "midipy_analysis_settings_signature"
)

results_are_current = (
    bool(analysis_payload)
    and stored_upload_signature == current_upload_signature
    and stored_settings_signature == current_settings_signature
)

with progress_placeholder:
    render_progress_stepper(
        active_step=4 if results_are_current else 2
    )

if analysis_payload and results_are_current:
    datasets = analysis_payload["datasets"]
    valid_names_by_role = st.session_state.get(
        "midipy_valid_names",
        {},
    )
    skipped_files_by_role = st.session_state.get(
        "midipy_skipped_files",
        {},
    )

    comparison_summary = pd.DataFrame()

    if comparison_enabled:
        comparison_summary = build_comparison_summary(
            datasets[ROLE_THERAPIST],
            datasets[ROLE_PARTICIPANT],
            selected_metrics,
        )

    export_tables = comparison_export_tables(
        datasets,
        comparison_summary,
    )
    excel_bytes = dataframe_to_excel_bytes(export_tables)
    csv_zip_bytes = dataframes_to_csv_zip(export_tables)

    total_processed = sum(
        len(names)
        for names in valid_names_by_role.values()
    )
    total_skipped = sum(
        len(items)
        for items in skipped_files_by_role.values()
    )

    st.markdown(
        f"""
        <div class="mp-results-banner" role="status">
            ✓ Analysis completed · {total_processed} file(s) processed ·
            {total_skipped} skipped
        </div>
        """,
        unsafe_allow_html=True,
    )

    results_heading, download_actions = st.columns(
        [3.7, 2],
        vertical_alignment="bottom",
    )

    with results_heading:
        st.markdown(
            """
            <h2 class="mp-section-title">4. Review results</h2>
            <p class="mp-section-copy">
                Review the comparison and the underlying performer data.
            </p>
            """,
            unsafe_allow_html=True,
        )

    with download_actions:
        excel_column, csv_column = st.columns(2)

        with excel_column:
            st.download_button(
                "Download Excel",
                data=excel_bytes,
                file_name=(
                    "MidiPy_Comparison.xlsx"
                    if comparison_enabled
                    else "MidiPy_Results.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

        with csv_column:
            st.download_button(
                "Download CSV",
                data=csv_zip_bytes,
                file_name=(
                    "MidiPy_Comparison_CSV.zip"
                    if comparison_enabled
                    else "MidiPy_CSV_Results.zip"
                ),
                mime="application/zip",
                use_container_width=True,
            )

    if comparison_enabled:
        therapist_valid = valid_names_by_role.get(
            ROLE_THERAPIST,
            [],
        )
        participant_valid = valid_names_by_role.get(
            ROLE_PARTICIPANT,
            [],
        )

        st.markdown(
            f"""
            <div class="mp-comparison-banner">
                <div>
                    <small>Music therapist</small>
                    <strong>{len(therapist_valid)} file(s) analyzed</strong>
                </div>
                <div>
                    <small>Participant</small>
                    <strong>{len(participant_valid)} file(s) analyzed</strong>
                </div>
                <div>
                    <small>Shared measures</small>
                    <strong>{len(comparison_summary)} available</strong>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.subheader("Comparison overview")
        st.caption(
            "Count measures are summed across files. Velocity and asynchrony "
            "measures are averaged. Differences are calculated as "
            "participant minus music therapist."
        )

        comparison_chart_column, comparison_table_column = st.columns(
            [1, 1.25],
            gap="large",
        )

        with comparison_chart_column:
            render_comparison_chart(comparison_summary)

        with comparison_table_column:
            if not comparison_summary.empty:
                display_comparison = comparison_summary.drop(
                    columns=["Metric key"],
                    errors="ignore",
                )
                st.dataframe(
                    display_comparison,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        ROLE_THERAPIST: st.column_config.NumberColumn(
                            ROLE_THERAPIST,
                            format="%.2f",
                        ),
                        ROLE_PARTICIPANT: st.column_config.NumberColumn(
                            ROLE_PARTICIPANT,
                            format="%.2f",
                        ),
                        "Difference": st.column_config.NumberColumn(
                            "Participant − therapist",
                            format="%.2f",
                        ),
                        "Relative difference (%)": (
                            st.column_config.NumberColumn(
                                "Relative difference",
                                format="%.1f%%",
                            )
                        ),
                    },
                )
            else:
                st.info(
                    "No shared whole-file or segment measures are available."
                )

        st.markdown(
            """
            <p class="mp-neutral-note">
                A higher or lower value is not automatically better. Interpretation
                depends on the measure, instrument, therapeutic goal, and session context.
            </p>
            """,
            unsafe_allow_html=True,
        )

        render_segment_comparison(
            datasets[ROLE_THERAPIST],
            datasets[ROLE_PARTICIPANT],
            selected_metrics,
        )

        st.subheader("Individual performer results")
        therapist_tab, participant_tab = st.tabs(
            [ROLE_THERAPIST, ROLE_PARTICIPANT]
        )

        with therapist_tab:
            render_dataset_details(
                role=ROLE_THERAPIST,
                dataset_results=datasets[ROLE_THERAPIST],
                valid_names=valid_names_by_role.get(
                    ROLE_THERAPIST,
                    [],
                ),
                skipped_files=skipped_files_by_role.get(
                    ROLE_THERAPIST,
                    [],
                ),
                include_feet=include_feet,
            )

        with participant_tab:
            render_dataset_details(
                role=ROLE_PARTICIPANT,
                dataset_results=datasets[ROLE_PARTICIPANT],
                valid_names=valid_names_by_role.get(
                    ROLE_PARTICIPANT,
                    [],
                ),
                skipped_files=skipped_files_by_role.get(
                    ROLE_PARTICIPANT,
                    [],
                ),
                include_feet=include_feet,
            )

    else:
        single_results = datasets[ROLE_SINGLE]

        overview_tab, data_tab, quality_tab = st.tabs(
            ["Overview", "Detailed data", "File quality"]
        )

        with overview_tab:
            render_dataset_summary(
                single_results,
                valid_names_by_role.get(ROLE_SINGLE, []),
                skipped_files_by_role.get(ROLE_SINGLE, []),
                include_feet,
            )

            result_options = list(single_results.keys())
            selected_result_name = st.radio(
                "Result set",
                options=result_options,
                horizontal=True,
                format_func=lambda name: (
                    "Whole-file results"
                    if name == "Whole_File_Results"
                    else "Segment results"
                ),
                key=f"result_set_single_{analysis_cycle}",
            )

            render_result_chart(
                single_results[selected_result_name],
                selected_result_name,
                key_prefix="single",
            )

        with data_tab:
            for result_name, dataframe in single_results.items():
                label = (
                    "Whole-file results"
                    if result_name == "Whole_File_Results"
                    else "Segment results"
                )

                with st.expander(
                    label,
                    expanded=len(single_results) == 1,
                ):
                    display_dataframe(dataframe)
                    st.caption(
                        f"{len(dataframe):,} row(s) × "
                        f"{len(dataframe.columns):,} column(s)"
                    )

        with quality_tab:
            quality_left, quality_right = st.columns(2)

            with quality_left:
                st.markdown("**Successfully analyzed**")
                for filename in valid_names_by_role.get(
                    ROLE_SINGLE,
                    [],
                ):
                    st.write(f"✓ {filename}")

            with quality_right:
                st.markdown("**Skipped during validation**")
                single_skipped = skipped_files_by_role.get(
                    ROLE_SINGLE,
                    [],
                )

                if single_skipped:
                    st.dataframe(
                        pd.DataFrame(
                            single_skipped,
                            columns=["File", "Reason"],
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )
                else:
                    st.write("✓ No files were skipped.")

else:
    st.info(
        "The results area will appear after the MIDI files have been analyzed."
    )


st.markdown(
    """
    <div class="mp-footer">
        MidiPy Analysis Studio · Clear, guided, and neutral performer comparison
    </div>
    """,
    unsafe_allow_html=True,
)
